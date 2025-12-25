from openpyxl import Workbook, load_workbook
path1=input("快输入你的源表格文件路径吧，记得复制全：")


wb0 = load_workbook(path1)
ws0 = wb0.active

wb1 = Workbook()
ws1 = wb1.acti
ws1["a1"] = "主要开课平台"
ws1["b1"] = "课程总数"
ws1["c1"] = "课程总数排名"
ws1["d1"] = "2017年课程数"
ws1["e1"] = "2017年排名"
ws1["f1"] = "2018年课程数"
ws1["g1"] = "2018年排名"
ws1["h1"] = "2020年课程数"
ws1["i1"] = "2020年排名"
ws1["j1"] = "2023年课程数"
ws1["k1"] = "2023年排名"
ws1["l1"] = "2025年课程数"
ws1["m1"] = "2025年排名"

maxrows0 = ws0.max_row


uniqueset1 = set()
for q in range(2, maxrows0 + 1):
    cell = ws0["g" + str(q)]
    if cell.value is not None and str(cell.value).strip() != "":
        uniqueset1.add(cell.value)

mylist = list(uniqueset1)
print(mylist)



platform_stats = {}
for platform in mylist:
    cnt0 = 0
    cnt2017 = 0
    cnt2018 = 0
    cnt2020 = 0
    cnt2023 = 0
    cnt2025 = 0

    for i in range(2, maxrows0 + 1):
        platform_cell = ws0["g" + str(i)]
        year_cell = ws0["b" + str(i)]

        if platform_cell.value == platform:
            cnt0 += 1

            if year_cell.value == 2017:
                cnt2017 += 1
            elif year_cell.value == 2018:
                cnt2018 += 1
            elif year_cell.value == 2020:
                cnt2020 += 1
            elif year_cell.value == 2023:
                cnt2023 += 1
            elif year_cell.value == 2025:
                cnt2025 += 1

    platform_stats[platform] = {
        'total': cnt0,
        '2017': cnt2017,
        '2018': cnt2018,
        '2020': cnt2020,
        '2023': cnt2023,
        '2025': cnt2025
    }

sorted_by_total = sorted(platform_stats.items(), key=lambda x: x[1]['total'], reverse=True)
total_rank = {platform: rank + 1 for rank, (platform, _) in enumerate(sorted_by_total)}


sorted_by_2017 = sorted(platform_stats.items(), key=lambda x: x[1]['2017'], reverse=True)
rank_2017 = {platform: rank + 1 for rank, (platform, _) in enumerate(sorted_by_2017)}

sorted_by_2018 = sorted(platform_stats.items(), key=lambda x: x[1]['2018'], reverse=True)
rank_2018 = {platform: rank + 1 for rank, (platform, _) in enumerate(sorted_by_2018)}

sorted_by_2020 = sorted(platform_stats.items(), key=lambda x: x[1]['2020'], reverse=True)
rank_2020 = {platform: rank + 1 for rank, (platform, _) in enumerate(sorted_by_2020)}

sorted_by_2023 = sorted(platform_stats.items(), key=lambda x: x[1]['2023'], reverse=True)
rank_2023 = {platform: rank + 1 for rank, (platform, _) in enumerate(sorted_by_2023)}

sorted_by_2025 = sorted(platform_stats.items(), key=lambda x: x[1]['2025'], reverse=True)
rank_2025 = {platform: rank + 1 for rank, (platform, _) in enumerate(sorted_by_2025)}


k = 2
for platform, stats in sorted_by_total:
    ws1["a" + str(k)] = platform
    ws1["b" + str(k)] = stats['total']
    ws1["c" + str(k)] = total_rank[platform]
    ws1["d" + str(k)] = stats['2017']
    ws1["e" + str(k)] = rank_2017[platform]
    ws1["f" + str(k)] = stats['2018']
    ws1["g" + str(k)] = rank_2018[platform]
    ws1["h" + str(k)] = stats['2020']
    ws1["i" + str(k)] = rank_2020[platform]
    ws1["j" + str(k)] = stats['2023']
    ws1["k" + str(k)] = rank_2023[platform]
    ws1["l" + str(k)] = stats['2025']
    ws1["m" + str(k)] = rank_2025[platform]
    k += 1

wb1.save("c:\\线上一流课程统计结果.xlsx")
print("结果已保存到: c:\\线上一流课程统计结果.xlsx")
print("ooooooooooooooooooooooooooooooooooooooover!!!!!")

