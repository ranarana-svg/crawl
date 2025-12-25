import openpyxl
import os
import re
from collections import Counter

# 获取桌面路径
desktop = os.path.join(os.path.expanduser('~'), 'Desktop')

# 文件路径
source_file = os.path.join(desktop, '高校一流课程建设数据.xlsx')
target_file = os.path.join(desktop, '线上一流课程统计结果.xlsx')

print("正在读取源文件...")
wb_source = openpyxl.load_workbook(source_file)
ws_source = wb_source.active

# 读取表头
headers = [cell.value for cell in ws_source[1]]

# 确定列索引
type_col = headers.index('课程类型') if '课程类型' in headers else 0
year_col = headers.index('年份') if '年份' in headers else 1
name_col = headers.index('课程名称') if '课程名称' in headers else 2
director_col = headers.index('课程负责人') if '课程负责人' in headers else 3
platform_col = headers.index('主要上线平台') if '主要上线平台' in headers else 6

# 读取所有数据（从第2行开始）
platforms = set()
directors = set()
year_counter = Counter()
total_count = 0

for row in ws_source.iter_rows(min_row=2, values_only=True):
    if row[0] is None:  # 跳过空行
        continue
    
    total_count += 1
    year = row[year_col]
    director = row[director_col]
    platform = row[platform_col]
    
    # 统计平台（非空）
    if platform and str(platform).strip():
        platforms.add(str(platform).strip())
    
    # 统计负责人（非空）
    if director and str(director).strip():
        directors.add(str(director).strip())
    
    # 统计年份
    if year:
        year_counter[year] += 1

print(f"总课程数: {total_count}")
print(f"平台数量: {len(platforms)}")
print(f"负责人数量: {len(directors)}")
print(f"年份统计: {dict(year_counter)}")

# 读取目标文件
print("\n正在读取目标文件...")
wb_target = openpyxl.load_workbook(target_file)
ws_target = wb_target.active

# 读取统计项目（第一行）
statistics_items = []
for cell in ws_target[1]:
    statistics_items.append(cell.value)

# 计算统计数据
results = {}
for idx, item in enumerate(statistics_items, start=1):
    if item is None:
        results[item] = 0
        continue
        
    item_str = str(item)
    
    # 根据位置和内容匹配
    # 列1: 主要上线平台
    if idx == 1:
        results[item] = len(platforms)
    # 列2: 课程负责人
    elif idx == 2:
        results[item] = len(directors)
    # 列3: 课程建设数量
    elif idx == 3:
        results[item] = total_count
    # 其他列：年份相关统计
    elif '课程数' in item_str:
        # 提取年份
        year_match = re.search(r'(\d{4})', item_str)
        if year_match:
            year = int(year_match.group(1))
            count = year_counter.get(year, 0)
            results[item] = count
        else:
            results[item] = 0
    else:
        # 占比统计（以年份开头但不是"课程数"）
        year_match = re.search(r'(\d{4})', item_str)
        if year_match:
            year = int(year_match.group(1))
            count = year_counter.get(year, 0)
            if total_count > 0:
                percentage = round(count / total_count * 100, 2)
            else:
                percentage = 0
            results[item] = f"{percentage}%"
        else:
            results[item] = 0

# 将结果写入目标文件（第二行）
print("\n正在写入统计结果...")
for col_idx, item in enumerate(statistics_items, start=1):
    value = results.get(item, 0)
    ws_target.cell(row=2, column=col_idx, value=value)
    print(f"  列{col_idx}: {item} = {value}")

# 保存文件
wb_target.save(target_file)
print(f"\n统计完成！结果已保存到: {target_file}")
